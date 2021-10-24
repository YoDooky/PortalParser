import re

import requests
from bs4 import BeautifulSoup
import lxml
from selenium import webdriver
from openpyxl import load_workbook

database_path = "C:/PyProject/PortalParser/database/users_info.xlsx"
workbook = load_workbook(filename=database_path)


def download_user_pages():
    url = 'https://portal.irkutskoil.ru/'
    target_url = 'https://portal.irkutskoil.ru/personal/?USER_ID='
    start_id = 6041
    end_id = 25152  # ID первого и последнего юзверя (6041 по 25152)
    login_route = 'login/'
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                      'Chrome/95.0.4638.54 Safari/537.36',
        'Origin': 'https://portal.irkutskoil.ru', 'Referer': url + login_route}
    s = requests.session()
    login_payload = {'login': '79526380590', 'pwd': '58r93j'}
    s.post(url + login_route, headers=headers, data=login_payload)  # проходим авторизацию
    for i in range(start_id, end_id):
        target_req = s.get(target_url + str(i))
        with open("all_users/user=" + str(i) + ".html", "w", encoding='utf-8') as file:
            file.write(target_req.text)


def get_data_from_source(user_id):
    with open("all_users/user=" + str(user_id) + ".html", encoding='utf-8') as file:
        src = file.read()
    soup = BeautifulSoup(src, "lxml")
    user_data = []
    # Если нет номера телефона то выходим из функции
    try:
        number = soup.find('div', text=re.compile('([Сс]отовый телефон)')).find_next().text
        if not re.search(r'\d', number):
            return
    except Exception as ex:
        return

    # Парсим ФИО
    try:
        user_name = soup.find('h1', {'style': 'margin-top: 0;'})
        user_data.append(user_name.text)
    except Exception as ex:
        return

    # Парсим должность
    try:
        user_post = soup.find('h3', {'style': 'margin-top: 10px;'})
        user_data.append(user_post.text)
    except Exception as ex:
        user_data.append('None')

    # Парсим сотовый телефон
    try:
        user_phone = soup.find('div', text=re.compile('([Сс]отовый телефон)')).find_next()
        user_data.append(user_phone.text)
    except Exception as ex:
        user_data.append('None')

    return user_data


def copy_to_excel(row_numb, users_data):
    sheet_name = 'users'
    workbook[sheet_name].cell(row=row_numb, column=1).value = users_data[0]
    workbook[sheet_name].cell(row=row_numb, column=2).value = users_data[1]
    workbook[sheet_name].cell(row=row_numb, column=3).value = users_data[2]
    workbook.save(database_path)
    pass


def main():
    row_numb = 0
    for i in range(10000, 10200):
        users_data = get_data_from_source(i)
        if not users_data:
            continue
        row_numb += 1
        copy_to_excel(row_numb, users_data)


main()
