# модуль для поиска данных в базе данных Excel и добавления их в массив
from openpyxl import load_workbook
from test_solving import clear_symbols

files_path = "C:/Prometei/"  # путь к папке со всеми файлами (драйвер хрома, база данных и т.п.)
workbook = load_workbook(filename=files_path + "SDO.xlsx")


# ищем лист с нужной темой (тема написана в 1 строке и столбце листа базы данных)
def find_course_sheet(course_name):
    for each in workbook.sheetnames:
        sheet = workbook[each]
        clear_course_name = clear_symbols(course_name)
        database_course_name = clear_symbols(str(sheet.cell(row=1, column=1).value))
        if clear_course_name == database_course_name:
            return sheet
        elif each == workbook.sheetnames[len(workbook.sheetnames) - 1]:
            print('Тема:', course_name,
                  'не найдена. Начинаю создавать ёбнутый лист со всеми возможными вопросами и ответами')
            # all_theme_list()
            all_sheet = all_theme_list()
            return all_sheet


def all_theme_list():
    sheet_name = "ALL"
    for each in workbook.sheetnames:  # цикл для удаления (в случае если есть) и создания листа со всеми вопросами
        if each == sheet_name:
            workbook.remove(workbook.get_sheet_by_name(each))
            workbook.create_sheet(sheet_name)
            break
        elif each == workbook.sheetnames[len(workbook.sheetnames) - 1]:
            workbook.create_sheet(sheet_name)
    question_column = []
    answer_column = []
    last_element = 2
    for each in workbook.sheetnames:
        sheet = workbook[each]
        row_numb = 2  # номер строки для итерации по строкам листа
        last_row = 'text'
        while last_row != 'None':  # заполняем массивы с вопросами и ответами непустыми значениями с текущего листа
            last_row = str(sheet.cell(row=row_numb, column=2).value)
            if last_row != 'None':
                question_column.append(str(sheet.cell(row=row_numb, column=2).value))
                answer_column.append(str(sheet.cell(row=row_numb, column=3).value))
            row_numb += 1
    for element in range(2, len(question_column) + 1):
        workbook[sheet_name].cell(row=last_element, column=2).value = question_column[element - 2]
        workbook[sheet_name].cell(row=last_element, column=3).value = answer_column[element - 2]
        last_element += 1
        # составляем массив вопросов и ответов из базы данных в нужной теме
    workbook.save("C:\Prometei\SDO.xlsx")
    return sheet_name


# функция для создания листа со всеми вопросами если не нашел лист по названию темы
def get_array_from_database():
    try:
        sheet = "ВСЕ ОТВЕТЫ ВСЕЛЕННОЙ"  # find_course_sheet()
    except Exception:
        print('Что-то пошло не так при поиске темы в функции "find_course_sheet"')
        return
    if sheet != 0:
        database_array = []  # массив с вопросами и соответствующими ответами
        answer_value = []  # массив с ответами (т.к. может быть несколько столбцов с ответами (васяткино ИЛИ))
        database_question_array = []
        database_answer_array = []
        for i in range(2, 100000):  # цикл для помещения всех вопросов в массив (ищет максимум в 100000 строк). Введи потом переменную которая считает количесво строк
            question_value = workbook[sheet].cell(row=i, column=2).value
            x = 3  # переменная для прохода по столбцам в цикле while
            answer_value = []  # очищаем массив с ответами чтобы не было дублей в "database_answer_array"
            while workbook[sheet].cell(row=i, column=x).value:
                answer_value.append(str(workbook[sheet].cell(row=i, column=x).value))
                x = x+1
            if question_value and answer_value:  # проверка пустые ли ячейки с впоросом и ответом, если пустые то
                # цикл заканчивается
                database_question_array.append(str(question_value))
                database_answer_array.append(answer_value)
            elif not question_value and not answer_value:
                break
        database_array.append(database_question_array)
        database_array.append(database_answer_array)
        workbook.close()
        return database_array
    else:
        workbook.close()
        return 0
